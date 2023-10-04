import openpyxl

# file-->workbook--->sheets-->Rows--->Cells
file = "Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row  # count number of rows in a excel sheet
colm = sheet.max_column  # count number of columns in a excel sheet

for r in range(1, rows + 1):
    for c in range(1, colm + 1):
        print(sheet.cell(r, c).value, end=" ")
    print()

