import openpyxl

#same data
"""file ="C:/Users/SARIKA/Desktop/Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "welcome"
workbook.save(file)"""

#different data
file="C:/Users/SARIKA/Desktop/Excel/Test1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

sheet.cell(1,1).value="hulk"
sheet.cell(1,2).value="50"
sheet.cell(2,1).value="thor"
sheet.cell(2,2).value="42"
sheet.cell(3,1).value="Groot"
sheet.cell(3,2).value="25"

workbook.save(file)
