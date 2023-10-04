import openpyxl
from openpyxl.styles import PatternFill
def read(filename,SheetName,RowNo,ColumnNo):
    file=filename
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(RowNo,ColumnNo).value

def write(FileName,SheetName,RowNo,ColumnNo,CellValue):
    file=FileName
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    sheet.cell(RowNo,ColumnNo).value=CellValue
    workbook.save(file)

def get_rows(FileName,SheetName):
    file=FileName
    woorkbook=openpyxl.load_workbook(file)
    sheet=woorkbook[SheetName]
    return(sheet.max_row)


def get_column(FileName,SheetName):
    file=FileName
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return(sheet.max_column)

def fillGreenColor(FileName,SheetName,RowNo,ColumnNo):
    file=FileName
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(RowNo,ColumnNo).value=greenFill
    workbook.save(file)

def fillRedColor(FileName,SheetName,RowNo,ColumnNo):
    file=FileName
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    redFill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(RowNo,ColumnNo).value=redFill
    workbook.save(file)


