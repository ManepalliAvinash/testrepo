"""import mysql.connector

con=mysql.connector.connect(host="localhost",port=3306,user="root",password="root",database="automation")
curs=con.cursor()
curs.execute("insert into tab1 values(1,'groot',25)")
curs.execute("insert into tab1 values(2,'hulk',20)")
curs.execute("insert into tab1 values(3,'thor',21)")
curs.execute("insert into tab1 values(4,'iron',11)")
curs.execute("insert into tab1 values(5,'Dragon',15)")
curs.execute("insert into tab1 values(6,'Nancy',29)")

curs.execute("select * from tab1")
for row in curs:
    print(row[0],row[1],row[2])"""

"""
import openpyxl

file="C:/Users/SARIKA/Desktop/Excel/Test2.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

sheet.cell(1,1).value="1"
sheet.cell(1,2).value="groot"
sheet.cell(1,3).value="25"
sheet.cell(2,1).value="2"
sheet.cell(2,2).value="hulk"
sheet.cell(2,3).value="18"
sheet.cell(3,1).value="3"
sheet.cell(3,2).value="thor"
sheet.cell(3,3).value="25"

workbook.save(file)

row=sheet.max_row
colmn=sheet.max_column

for r in range(1,row+1):
    for c in range(1,colmn+1):
        print(sheet.cell(r,c).value,end=" ")
    print()
"""
"""from selenium import webdriver

def chrome_setup():
    ops=webdriver.ChromeOptions()
    ops.headless=True
    driver=webdriver.Chrome(options=ops)
    return driver

driver=chrome_setup()
driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()
print(driver.current_url)
print(driver.title)

driver.close()"""


"""from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://demo.nopcommerce.com/")

driver.maximize_window()
cookies=driver.get_cookies()
print("length of cookies before:",len(cookies))
for cookie in cookies:
    print(cookie)

driver.add_cookie({"name":'Hulk',"value":"1234"})
cookies=driver.get_cookies()
print("lenght of cookies after adding cookie: ",len(cookies))

driver.delete_cookie("Hulk")
cookies=driver.get_cookies()
print("lenght of cookies after deleting the added cookie: ",len(cookies))

driver.delete_all_cookies()
cookies=driver.get_cookies()
print("lenght of cookies after deleting the added cookie: ",len(cookies))

time.sleep(10)
driver.close()
"""

"""from selenium import webdriver
import time
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()
#reg_link=Keys.CONTROL+Keys.RETURN
#driver.find_element(By.XPATH,"//a[@class='ico-register']").send_keys(reg_link)
driver.switch_to.new_window("window")
driver.get("https://www.naukri.com/mnjuser/homepage")
time.sleep(10)
driver.quit()"""

from selenium import webdriver
from selenium.webdriver.common.by import By

# no of rows

driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
rows=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
print(len(rows))

#no of column

columns=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr[1]/th")
print(len(columns))

# read specfic row and column
spc_value=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[3]/td[1]").text
if spc_value=="Learn Java":
    print(spc_value)
    print("spc value matched")

# read all values from rows and columns
for r in range(2,len(rows)+1):
    for c in range(1,len(columns)+1):
        data=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
        print(data,end=" ")
    print()
print("completed")
# to print book name if autor is matched

for r in range(2,len(rows)+1):
    Author_name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[2]").text
    if Author_name=="Mukesh":
        Price=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[4]").text
        Book_Name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[1]").text
        print(Price,Book_Name)
driver.close()


