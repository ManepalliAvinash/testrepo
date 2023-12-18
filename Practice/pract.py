"""import openpyxl
file="C:/Users/SARIKA/Desktop/Excel/Test1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
colm= sheet.max_column

for i in range(1,rows+1):
    for j in range(1,colm+1):
        print(sheet.cell(i,j).value,end=" ")
    print(" ")"""
import time

from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select

"""create ="create table data(sno int(2),name varchar(20),marks int(2))"
insert1="insert into data values(1,'hulk',22)"
insert2="insert into data values(2,'groot',21)"
insert3="insert into data values(3,'donna',33)"
insert4="insert into data values(4,'teddy',22)"
read="select * from data"

import mysql.connector
con=mysql.connector.connect(host="localhost",port=3306,user="root",password="root",database="automation")
curs=con.cursor()
curs.execute(create)
curs.execute(insert1)
curs.execute(insert4)
curs.execute(insert3)
curs.execute(insert2)
curs.execute(read)
for rows in curs:
    print(rows[0],rows[1],rows[2])"""



"""driver=webdriver.Chrome()
driver.get("https://the-internet.herokuapp.com/javascript_alerts")
driver.maximize_window()
driver.find_element(By.XPATH,"//button[@onclick='jsAlert()']").click()
alert_key=driver.switch_to.alert
print(alert_key.text)
time.sleep(5)
alert_key.accept()
result=driver.find_element(By.XPATH,"//p[@id='result']").text
print(result)"""

"""driver=webdriver.Chrome()
driver.get("https://the-internet.herokuapp.com/javascript_alerts")
driver.maximize_window()
driver.find_element(By.XPATH,"//button[@onclick='jsPrompt()']']").click()
alert_key=driver.switch_to.alert
print(alert_key.text)
time.sleep(5)
alert_key.accept()
result=driver.find_element(By.XPATH,"//p[@id='result']").text
print(result)"""


"""driver=webdriver.Chrome()
driver.get("https://the-internet.herokuapp.com/javascript_alerts")
driver.maximize_window()
driver.find_element(By.XPATH,"//button[@onclick='jsPrompt()']").click()
alert_key=driver.switch_to.alert
alert_key.send_keys("Hulk welcome to my world")
print(alert_key.text)
time.sleep(5)
alert_key.accept()
result=driver.find_element(By.XPATH,"//p[@id='result']").text
print(result)"""

"""driver=webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Frames.html")
driver.maximize_window()
driver.find_element(By.XPATH,"//a[normalize-space()='Iframe with in an Iframe']").click()
outer=driver.find_element(By.XPATH,"//*[@id='Multiple']/iframe")
driver.switch_to.frame(outer)

inner=driver.find_element(By.CSS_SELECTOR,"iframe[src='SingleFrame.html']")
driver.switch_to.frame(inner)

driver.find_element(By.XPATH,"//input[@type='text']").send_keys("groot")
time.sleep(10)"""

"""from selenium import webdriver
from selenium.webdriver.common.by import By
driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
rows=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
print(len(rows))
num_rows=len(rows)

column=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr/th")
print(len(column))
num_colm=len(column)
for i in range(2,num_rows+1):
    for j in range(1,num_colm+1):
        data=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td["+str(j)+"]").text
        print(data,end=" ")
    print()

for i in range(2,num_rows+1):
    author_name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[2]").text
    if author_name=="Amod":
        book_name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[1]").text
        Subject=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[3]").text
        price=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[4]").text
        print(book_name,author_name,Subject,price)

cntry_elm=driver.find_element(By.XPATH,"//select[@id='country']")
elm=Select(cntry_elm)
elm.select_by_visible_text("India")
time.sleep(10)"""

Date="27"
Year="2001"
Month="December"

"""from selenium import webdriver
from selenium.webdriver.common.by import By
driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
driver.find_element(By.XPATH,"//input[@id='datepicker']").click()

while True:
    month=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-month']").text
    year=driver.find_element(By.XPATH,"//span[@class='ui-datepicker-year']").text
    if Month==month and Year==year:
        break
    else:
        driver.find_element(By.XPATH,"//*[@id='ui-datepicker-div']/div/a[1]/span").click()

dates=driver.find_elements(By.XPATH,"//table[@class='ui-datepicker-calendar']/tbody/tr/td/a")
for elm in dates:
    if elm.text==Date:
        elm.click()
        break
time.sleep(10)"""
"""from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
driver.maximize_window()
driver.find_element(By.XPATH,"//input[@id='dob']").click()
month_elm=driver.find_element(By.XPATH,"//select[@aria-label='Select month']")
mon_elm=Select(month_elm)
mon_elm.select_by_visible_text("Dec")

year_elm=driver.find_element(By.XPATH,"//select[@aria-label='Select year']")
yr_elm=Select(year_elm)
yr_elm.select_by_value("2001")

date=driver.find_elements(By.XPATH,"//table[@class='ui-datepicker-calendar']/tbody/tr/td/a")
for elm in date:
    if elm.text==Date:
        elm.click()
        break
time.sleep(10)"""

"""from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
dbl_clk=driver.find_element(By.XPATH,"//*[@id='HTML10']/div[1]/button")
act=ActionChains(driver)
act.double_click(dbl_clk).perform()
time.sleep(10)

drag=driver.find_element(By.XPATH,"//p[normalize-space()='Drag me to my target']")
drop=driver.find_element(By.XPATH,"//div[@id='droppable']")

act=ActionChains(driver)
act.drag_and_drop(drag,drop).perform()
time.sleep(10)


driver=webdriver.Chrome()
driver.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
driver.maximize_window()
rgt_clk=driver.find_element(By.XPATH,"/html/body/div[1]/section/div/div/div/p/span")
act=ActionChains(driver)
act.context_click(rgt_clk).perform()
time.sleep(5)
driver.find_element(By.XPATH,"/html/body/ul/li[4]").click()
alrt_eml=driver.switch_to.alert
alrt_eml.accept()
time.sleep(10)
"""



"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
slider=driver.find_element(By.XPATH,"//span[@class='ui-slider-handle ui-corner-all ui-state-default']")
print("Before slider location: ",slider.location)

act=ActionChains(driver)
act.drag_and_drop_by_offset(slider,200,0).perform()
print("After slider location:",slider.location)
time.sleep(20)"""

"""driver=webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Slider.html")
driver.maximize_window()
sld=driver.find_element(By.XPATH,"//a[@class='ui-slider-handle ui-state-default ui-corner-all']")
print("slider loction before: ",sld.location)

act=ActionChains(driver)
act.drag_and_drop_by_offset(sld,180,0).perform()
print("after ald location: ",sld.location)
time.sleep(10)"""


"""from selenium import webdriver
from selenium.webdriver.common.by import By
import os
def chromeSetup():
    get_cwd=os.getcwd()
    preferences={"download:default_dirctory":get_cwd}
    opts=webdriver.ChromeOptions()
    opts.add_experimental_option("prefs",preferences)
    driver=webdriver.Chrome(options=opts)
    return driver

driver=chromeSetup()
driver.get("https://file-examples.com/index.php/sample-documents-download/sample-doc-download/")
driver.maximize_window()
driver.find_element(By.XPATH,"//tbody/tr[1]/td[5]/a[1]").click()
time.sleep(10)"""

"""from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
driver.maximize_window()
driver.find_element(By.XPATH,"//*[@id='select2-billing_country-container']").click()
ctr_elm=driver.find_elements(By.XPATH,"//*[@id='select2-billing_country-results']/li")
print(len(ctr_elm))

for ctr in ctr_elm:
    if ctr.text == "India":
        ctr.click()
        break
time.sleep(10)"""




"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
cookies=driver.get_cookies()
print(len(cookies),"before add")

driver.add_cookie({"name":"groot","value":"1234"})
cookie=driver.get_cookies()
print(cookie)
print(len(cookie),"after add")
cok=driver.get_cookie("groot")
print(cok)"""
"""import os
driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
getcwd=os.getcwd()
driver.save_screenshot(getcwd+".\\home1.png")
print(getcwd)

driver.get_screenshot_as_file(getcwd+".\\groot.png")"""

"""import mysql.connector
read="select * from data"
insert1="insert into data values(1,'hulk',22)"
insert2="insert into data values(2,'groot',21)"
insert3="insert into data values(3,'donna',33)"
insert4="insert into data values(4,'teddy',22)"
read="select * from data"

import mysql.connector
con=mysql.connector.connect(host="localhost",port=3306,user="root",password="root",database="automation")
curs=con.cursor()
curs.execute(insert1)
curs.execute(insert4)
curs.execute(insert3)
curs.execute(insert2)
curs.execute(read)
for rows in curs:
    print(rows[0])"""


"""from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys

driver=webdriver.Chrome()
driver.get("https://text-compare.com/")
driver.maximize_window()
input1=driver.find_element(By.XPATH,"//textarea[@id='inputText1']")
input2=driver.find_element(By.XPATH,"//textarea[@id='inputText2']")
input1.send_keys("Hello groot Welcome to my world")

act=ActionChains(driver)
act.key_down(Keys.CONTROL).send_keys("A").key_up(Keys.CONTROL).perform()
act.key_down(Keys.CONTROL).send_keys("C").key_up(Keys.CONTROL).perform()
act.key_down(Keys.TAB).key_up(Keys.TAB).perform()
act.key_down(Keys.CONTROL).send_keys("V").key_up(Keys.CONTROL).perform()

time.sleep(10)
driver.close()"""

"""from selenium import webdriver
from selenium.webdriver.common.by import By

def headleass_mode():
    ops=webdriver.ChromeOptions()
    ops.headless=True
    driver=webdriver.Chrome(options=ops)
    return driver

driver=headleass_mode()
driver.get("https://testautomationpractice.blogspot.com/")
title=driver.title
print(title)
url=driver.current_url
print(url)
print(driver.current_window_handle)"""

"""import openpyxl
files="C:/Users/SARIKA/Desktop/Excel/Fixed_Deposit.xlsx"
workbook=openpyxl.load_workbook(files)
sheet=workbook["Sheet1"]
rows=sheet.max_row
colm=sheet.max_column
for r in range(2,rows+1):
    for c in range(1,colm):
        print(sheet.cell(r,c).value,end=" ")
    print()"""



"""def chrome_setup():
    ops=webdriver.ChromeOptions()
    ops.headless = True
    driver=webdriver.Chrome(options=ops)
    return driver

driver=chrome_setup()
driver.get("https://smartinternz.com/student-login")
print(driver.current_url)
print(driver.page_source)
print(driver.current_window_handle)
print(driver.title)
driver.switch_to.new_window("Tab")
ins=driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
print(driver.current_url)
print(driver.page_source)
print(driver.current_window_handle)
print(driver.title)"""

"""from selenium import webdriver
from selenium.webdriver.common.by import By
import mysql.connector

driver=webdriver.Chrome()
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
driver.maximize_window()
driver.implicitly_wait(10)


con=mysql.connector.connect(host="localhost", port=3306, user="root", password="root", database="automation")
curs=con.cursor()
curs.execute("select * from caldata")
for row in curs:
    princ=row[0]
    roi=row[1]
    period1=row[2]
    period2=row[3]
    freq=row[4]
    expec_mtry_value=row[5]

    driver.find_element(By.XPATH,"//input[@id='principal']").send_keys(princ)
    driver.find_element(By.XPATH,"//input[@id='interest']").send_keys(roi)
    driver.find_element(By.XPATH,"//input[@id='tenure']").send_keys(period1)
    day_elm=driver.find_element(By.XPATH,"//select[@id='tenurePeriod']")
    days_elm=Select(day_elm)
    days_elm.select_by_visible_text(period2)
    freq_elm=driver.find_element(By.XPATH,"//select[@id='frequency']")
    freqs_elm=Select(freq_elm)
    freqs_elm.select_by_visible_text(freq)
    driver.find_element(By.XPATH, '//*[@id="fdMatVal"]/div[2]/a[1]').click()
    mtry_value=driver.find_element(By.XPATH,"//span[@id='resp_matval']/strong").text

    if float(expec_mtry_value)==float(mtry_value):
        print("Test passed")
    else:
        print("Test Failed")

    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[2]").click()
    time.sleep(5)


driver.close()"""

