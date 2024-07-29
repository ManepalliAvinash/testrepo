"""import mysql.connector
input1="insert into test1 values (1,'Avinash',3)"
input2="insert into test1 values(2,'ganesh',9)"
input3="insert into test1(sno,sname,tests_passed) values(3,'ravi',4)"
input4="insert into test1 values(4,'jenny',6)"
read_data="select * from test1"
connect=mysql.connector.connect(host="localhost",port=3306,user="root",password="root",database="testing")
cursor=connect.cursor()
cursor.execute(input1)
cursor.execute(input2)
cursor.execute(input3)
cursor.execute(input4)
cursor.execute(read_data)
for row in cursor:
    value1=row[0]
    value2=row[1]
    value3=row[2]
    print(value1,value2,value3)
connect.commit()
connect.close()"""

"""import mysql.connector

con=mysql.connector.connect(host="localhost",username="root",password="root",port=3306,database="automation")
curs=con.cursor()
curs.execute("drop table data8")
curs.execute("create table data8(sno int(2),sname varchar(20),marks int(2))")
curs.execute("insert into data8 values(1,'groot',22)")
curs.execute("insert into data8 values(2,'hulk',20)")
curs.execute("insert into data8 values(3,'witch',25)")
curs.execute("insert into data8 values(4,'ninja',20)")
con.commit()
con.close()
con.reconnect()
curs.execute("select * from data8")
for row in curs:
    print(row)"""

"""import openpyxl

file="C:/Users/SARIKA/Desktop/Excel/Fixed_Deposit.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row
column=sheet.max_column
print(rows)
print(column)
for r in range(2,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end=" ")
    print("")"""







