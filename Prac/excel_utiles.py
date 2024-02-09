import openpyxl

def num_of_rows(file,sheet_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_no]
    rows=sheet.max_row
    return rows

def num_of_column(file,sheet_no):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_no]
    column=sheet.max_column
    return column

def read_value(file,sheet_no,row,column):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_no]
    value=sheet.cell(row,column).value
    return value

def write_value(file,sheet_no,row,column,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheet_no]
    sheet.cell(row,column).value=data
    workbook.save(file)
file="C:/Users/SARIKA/Desktop/Excel/Fixed_Deposit.xlsx"
rows=num_of_rows(file,"Sheet1")
print(type(rows))