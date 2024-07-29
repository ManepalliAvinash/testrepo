import time

from selenium import webdriver
from selenium.webdriver import ActionChains,Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

"""import requests
driver=webdriver.Chrome()
driver.get("http://www.deadlinkcity.com/")
driver.maximize_window()
links=driver.find_elements(By.TAG_NAME, "a")
count=0
for link in links:
    url=link.get_attribute("href")

    try:
        response=requests.head(url)
    except:
        None
    if response.status_code>400:
        print("broken link")
        count+=1
    else:
        print("link is working")

driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
#driver.find_element(By.XPATH,"//button[@onclick='myFunctionAlert()']").click()"""
"""driver.find_element(By.XPATH,"//button[@onclick='myFunctionConfirm()']").click()
alert_elm=driver.switch_to.alert
alert_elm.accept()
text=driver.find_element(By.XPATH,"//*[@id='demo']").text
print(text)
driver.find_element(By.XPATH,"//button[@onclick='myFunctionConfirm()']").click()
alert_elm=driver.switch_to.alert
alert_elm.dismiss()
text=driver.find_element(By.XPATH,"//*[@id='demo']").text
print(text)
time.sleep(10)"""
"""driver.find_element(By.XPATH,"//button[@onclick='myFunctionPrompt()']").click()
alert=driver.switch_to.alert
alert.send_keys("Groot")
alert.accept()
text=driver.find_element(By.XPATH,"//p[@id='demo']").text
print(text)
driver.quit()"""

"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
driver.switch_to.frame("frame-one796456169")
driver.find_element(By.XPATH,"//input[@id='RESULT_TextField-0']").send_keys("Groot")"""
"""clickable=driver.find_element(By.XPATH,"//input[@id='RESULT_RadioButton-1_0']")
print(clickable.is_enabled())
clickable=driver.find_element(By.XPATH,"//input[@id='RESULT_RadioButton-1_0']").click()"""

"""driver.find_element(By.XPATH,"//*[@id='RESULT_RadioButton-1_0']").click()
driver.find_element(By.XPATH,"//div[@class='fauxborder-left main-fauxborder-left']").click()
year_elm=driver.find_element(By.XPATH,"//*[@id='ui-datepicker-div']/div/div/select")
year=Select(year_elm)
year.select_by_value("2001")
for i in range(12):
    month = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span").text
    if month == "December":
        break
    else:
        driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/a[2]").click()
date=driver.find_elements(By.XPATH,"//*[@id='ui-datepicker-div']/table/tbody/tr/td/a")
for i in date:
    if i.text=="27":
        i.click()
        break
job_elm=driver.find_element(By.XPATH,"//*[@id='RESULT_RadioButton-3']")
job=Select(job_elm)
job.select_by_visible_text("Automation Engineer")"""
"""time.sleep(10)
driver.quit()"""

"""driver=webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Frames.html")
driver.maximize_window()
driver.switch_to.frame("SingleFrame")
driver.find_element(By.XPATH,"//input[@type='text']").send_keys("welcome")
time.sleep(5)
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Frames.html")
driver.maximize_window()
driver.find_element(By.XPATH,"//a[normalize-space()='Iframe with in an Iframe']")
Multipleframe=driver.find_element(By.CSS_SELECTOR,"iframe[src='MultipleFrames.html']")
driver.switch_to.frame(Multipleframe)
Singleframe=driver.find_element(By.CSS_SELECTOR,"iframe[src='SingleFrame.html']")
driver.switch_to.frame(Singleframe)
driver.find_element(By.XPATH,"//input[@type='text']").send_keys("Hulk")

time.sleep(5)
driver.close()"""
"""driver=webdriver.Chrome()
driver.get("https://demo.nopcommerce.com/login?returnUrl=%2F")
driver.maximize_window()
driver.get("https://amazon.com/")
time.sleep(5)
driver.back()
time.sleep(5)
driver.forward()
time.sleep(5)
driver.refresh()
time.sleep(5)
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://demo.automationtesting.in/Windows.html")
driver.maximize_window()
driver.switch_to.new_window("tab")
driver.get("https://amazon.com/")
driver.switch_to.new_window("tab")
driver.get("https://demo.nopcommerce.com/login?returnUrl=%2F")
print(driver.title)
print(driver.current_window_handle)

window_ids=driver.window_handles
print(window_ids)

time.sleep(10)
driver.switch_to.window(window_ids[1])
print(driver.title)
time.sleep(5)
driver.switch_to.window(window_ids[0])
print(driver.title)
time.sleep(5)
driver.switch_to.window(window_ids[2])
print(driver.title)
time.sleep(5)
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
rows=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
len_rows=len(rows)
print("No of rows: ",rows)
columns=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr/th")
len_column=len(columns)
print("No of columns: ",columns)

value_by_row=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[2]/td[1]")
if value_by_row.text=="Learn Selenium":
    print("value matched")
else:
    print("value not matched")

for r in range(2,len_rows+1):
    for c in range(1,len_column+1):
        values=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td["+str(c)+"]")
        print(values.text,end="   ")
    print()



for r in range(2,len_rows+1):
    Author_Name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[2]").text
    if Author_Name=="Amit":
        Book_name=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[1]").text
        Subject=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[3]").text
        price=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(r)+"]/td[4]").text
        print(Book_name,"  ",Subject,"  ",price)


driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
driver.find_element(By.XPATH,"//input[@id='datepicker']").click()

date="27"
month="December"
year="2001"


while True:
    Year = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[2]").text
    Month = driver.find_element(By.XPATH, "//*[@id='ui-datepicker-div']/div/div/span[1]").text
    if month==Month and year==Year:
        break
    else:
        driver.find_element(By.XPATH,"//a[@title='Prev']").click()      #prev
        #driver.find_element(By.XPATH,"//a[@title='Next']").click()

Dates=driver.find_elements(By.XPATH,"//table[@class='ui-datepicker-calendar']/tbody/tr/td/a")
for Date in Dates:
    if date==Date.text:
        Date.click()
        break

time.sleep(10)
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
driver.maximize_window()

driver.find_element(By.XPATH,"//*[@id='dob']").click()
time.sleep(10)
month_elm=driver.find_element(By.XPATH,"//select[@aria-label='Select month']")
Month=Select(month_elm)
Month.select_by_visible_text("Dec")

year_elm=driver.find_element(By.XPATH,"//select[@aria-label='Select year']")
Year=Select(year_elm)
Year.select_by_visible_text("2001")

Dates=driver.find_elements(By.XPATH,"//table[@class='ui-datepicker-calendar']/tbody/tr/td/a")
for Date in Dates:
    if Date.text=="27":
        Date.click()
        break
driver.find_element(By.XPATH,"//button[normalize-space()='Done']").click()

time.sleep(10)
driver.close()"""

#driver=webdriver.Chrome()
#driver.get("https://testautomationpractice.blogspot.com/")
#driver.maximize_window()
"""dbl_clk=driver.find_element(By.XPATH,"//button[normalize-space()='Copy Text']")
act=ActionChains(driver)
act.double_click(dbl_clk).perform()
time.sleep(5)
text=driver.find_element(By.XPATH,"//*[@id='HTML10']/div[1]/p").text
print(text)"""
"""drag=driver.find_element(By.XPATH,"//p[normalize-space()='Drag me to my target']")
drop=driver.find_element(By.XPATH,"//div[@id='droppable']")
act=ActionChains(driver)
act.drag_and_drop(drag,drop).perform()
time.sleep(5)
text=driver.find_element(By.XPATH,"//*[@id='droppable']/p").text
print(text)"""
"""slider=driver.find_element(By.XPATH,"//*[@id='slider']/span")
print(slider.location)
act=ActionChains(driver)
act.drag_and_drop_by_offset(slider,80,0).perform()
time.sleep(10)
print(slider.location)"""
#driver.get("https://demo.automationtesting.in/Slider.html")
"""driver.maximize_window()
slider=driver.find_element(By.XPATH,"//a[@class='ui-slider-handle ui-state-default ui-corner-all']")
print(slider.location)
act=ActionChains(driver)
act.drag_and_drop_by_offset(slider,100,0).perform()
time.sleep(5)
print(slider.location)"""
"""driver.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
driver.maximize_window()
right_clk=driver.find_element(By.XPATH,"//span[@class='context-menu-one btn btn-neutral']")
act=ActionChains(driver)
act.context_click(right_clk).perform()
time.sleep(3)
driver.find_element(By.XPATH,"/html/body/ul/li[3]").click()
time.sleep(5)
alert=driver.switch_to.alert
print(alert.text)
alert.accept()
time.sleep(5)"""
#driver.close(),

"""driver=webdriver.Chrome()
driver.get("https://text-compare.com/")
driver.maximize_window()
driver.find_element(By.XPATH,"//textarea[@id='inputText1']").send_keys("Hello welcome to my world")

act=ActionChains(driver)
act.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
act.key_down(Keys.CONTROL).send_keys("c").key_up(Keys.CONTROL).perform()
act.key_down(Keys.TAB).key_up(Keys.TAB).perform()
act.key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
driver.find_element(By.XPATH,"//div[@class='compareButtonText']").click()
time.sleep(5)
text=driver.find_element(By.XPATH,"/html/body/div[2]/span").text
print(text)
driver.close()"""
"""import os
get_cwd=os.getcwd()
def chrome_setup():
    preferences={"download.default_directory": get_cwd}
    options=webdriver.ChromeOptions()
    options.add_experimental_option("prefs",preferences)
    driver=webdriver.Chrome(options=options)
    return driver

driver=chrome_setup()
driver.get("https://file-examples.com/index.php/sample-documents-download/sample-doc-download/")
driver.find_element(By.XPATH,"//tbody/tr[1]/td[5]/a[1]").click()
time.sleep(20)
driver.close()"""


"""import os
get_cwd=os.getcwd()
def chrome_setup():
    preferences={"download:default_directory":get_cwd}
    options=webdriver.ChromeOptions()
    options.add_experimental_option("prefs",preferences)
    driver=webdriver.Chrome(options=options)
    return driver

driver=chrome_setup()
driver.get("https://file-examples.com/index.php/sample-documents-download/sample-doc-download/")
driver.maximize_window()
driver.find_element(By.XPATH,"//body[1]/div[1]/main[1]/section[1]/div[1]/div[2]/div[1]/div[1]/table[1]/tbody[1]/tr[3]/td[5]/a[1]").click()
time.sleep(30)
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://www.dummyticket.com/dummy-ticket-for-visa-application/")
driver.maximize_window()
driver.find_element(By.XPATH,"//span[@id='select2-billing_country-container']").click()
countries=driver.find_elements(By.XPATH,"//*[@id='select2-billing_country-results']/li")
for country in countries:
    if country.text=="India":
        country.click()
        break
time.sleep(5)

driver.find_element(By.XPATH,"//*[@id='select2-billing_state-container']").click()
States=driver.find_elements(By.XPATH,"//*[@id='select2-billing_state-results']/li")
for state in States:
    if state.text=="Andhra Pradesh":
        state.click()
        break
time.sleep(5)
driver.close()"""

"""def chromesetup():
    options=webdriver.ChromeOptions()
    options.add_argument('--headless')
    driver=webdriver.Chrome(options=options)
    return driver

driver=chromesetup()
driver.get("https://testautomationpractice.blogspot.com/")
driver.find_element(By.XPATH,"//input[@id='name']").send_keys("Groot")
text=driver.find_element(By.XPATH,"//input[@id='name']").get_attribute("value")
print(text)
driver.quit()"""
"""import os
get_cwd=os.getcwd()
driver=webdriver.Chrome()
driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()
driver.save_screenshot(get_cwd+"//testautomation.png")
driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://demo.nopcommerce.com/")
driver.maximize_window()
cookies=driver.get_cookies()
print("Before: ",len(cookies))
print(cookies)

driver.add_cookie({"name":"Mycookie","value":"12345","date":"08"})
cookies=driver.get_cookies()
print("after:",len(cookies))
print(cookies)
driver.close()
"""
"""import openpyxl
file="C:/Users/SARIKA/Desktop/Excel/Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
rows=sheet.max_row
columns=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,columns+1):
        cell=sheet.cell(r,c).value
        print(cell,end=" ")
    print()"""
"""import openpyxl
import excel_utiles
driver=webdriver.Chrome()
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
driver.maximize_window()
file="C:/Users/SARIKA/Desktop/Excel/Fixed_Deposit.xlsx"
rows=excel_utiles.num_of_rows(file,"Sheet1")
for r in range(2,rows+1):
    princ=excel_utiles.read_value(file,"Sheet1",r,1)
    roi=excel_utiles.read_value(file,"Sheet1",r,2)
    period1=excel_utiles.read_value(file,"Sheet1",r,3)
    period2=excel_utiles.read_value(file,"Sheet1",r,4)
    freq=excel_utiles.read_value(file,"Sheet1",r,5)
    manturity_value=excel_utiles.read_value(file,"Sheet1",r,6)

    driver.find_element(By.XPATH,"//input[@id='principal']").send_keys(princ)
    driver.find_element(By.XPATH,"//input[@id='interest']").send_keys(roi)
    driver.find_element(By.XPATH,"//input[@id='tenure']").send_keys(period1)
    period2_elm=driver.find_element(By.XPATH,"//select[@id='tenurePeriod']")
    period2_opt=Select(period2_elm)
    period2_opt.select_by_visible_text(period2)
    freq_elm=driver.find_element(By.XPATH,"//select[@id='frequency']")
    freq_elms=Select(freq_elm)
    freq_elms.select_by_visible_text(freq)

    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[1]").click()
    actual_value=driver.find_element(By.XPATH,"//span[@id='resp_matval']/strong")
    if manturity_value==actual_value.text:
        print("Test Passed")
    else:
        print("Test Failed")
    time.sleep(10)
    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[2]").click()

driver.close()"""

"""import mysql.connector

connect=mysql.connector.connect(host="localhost",user="root",password="root",port=3306,database="automation")
curs=connect.cursor()
read="select * from caldata"
curs.execute(read)
driver = webdriver.Chrome()
driver.get(
        "https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
driver.maximize_window()
driver.implicitly_wait(2.0)
for row in curs:
    princ = row[0]
    roi = row[1]
    period1 =row[2]
    period2 = row[3]
    freq = row[4]
    expected_value = row[5]

    driver.find_element(By.XPATH, "//input[@id='principal']").send_keys(princ)
    driver.find_element(By.XPATH, "//input[@id='interest']").send_keys(roi)
    driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys(period1)
    period2_elm = driver.find_element(By.XPATH, "//select[@id='tenurePeriod']")
    period2_opt = Select(period2_elm)
    period2_opt.select_by_visible_text(period2)
    freq_elm = driver.find_element(By.XPATH, "//select[@id='frequency']")
    freq_elms = Select(freq_elm)
    freq_elms.select_by_visible_text(freq)

    driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]").click()
    actual_value = driver.find_element(By.XPATH, "//span[@id='resp_matval']/strong")
    if expected_value==actual_value.text:
        print("Test passed")
    else:
        print("Test failed")

    driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[2]").click()
    time.sleep(5)

driver.close()"""

"""driver=webdriver.Chrome()
driver.get("https://www.moneycontrol.com/fixed-income/calculator/state-bank-of-india-sbi/fixed-deposit-calculator-SBI-BSB001.html?classic=true")
driver.maximize_window()
driver.find_element(By.XPATH, "//input[@id='principal']").send_keys("10000")
driver.find_element(By.XPATH, "//input[@id='interest']").send_keys("12")
driver.find_element(By.XPATH, "//input[@id='tenure']").send_keys("3")
period2_elm = driver.find_element(By.XPATH, "//select[@id='tenurePeriod']")
period2_opt = Select(period2_elm)
period2_opt.select_by_visible_text("year(s)")
freq_elm = driver.find_element(By.XPATH, "//select[@id='frequency']")
freq_elms = Select(freq_elm)
freq_elms.select_by_visible_text("Simple Interest")
expected_value="13600.00"
driver.find_element(By.XPATH, "//*[@id='fdMatVal']/div[2]/a[1]").click()
actual_value = driver.find_element(By.XPATH, "//span[@id='resp_matval']/strong")
if expected_value==actual_value.text:
    print("Test passed")
else:
    print("Test failed")

driver.find_element(By.XPATH,"//*[@id='fdMatVal']/div[2]/a[2]").click()
time.sleep(5)
driver.close()"""

num=int(input("enter number:" ))
if num>1:
    for i in range(2,num):
        if (num%i==0):
            print("not prime")
            break
    else:
        print("prime")
else:
    print("num less than 1")